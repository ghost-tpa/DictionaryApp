# python 
# Created by Tuan Anh Phan on 03.08.2023
from openpyxl import open as openex
from os import getcwd, path
from my_crypto import newCipher, newHash

src_path = path.join(getcwd(), "src")
passwd_path = path.join(src_path, "passwd.txt")
passwd_path_test = path.join(src_path, "passwd_test.txt")

def main():
    pass

def read_data_from_excel(excel_file_name: str):
    path_exel_file = path.join(src_path, excel_file_name)
    sheet_name = "Sheet1"
    wb = openex(path_exel_file)
    sheet_obj = wb[sheet_name]
    data = []
    for row in sheet_obj.values:
        for value in row:
            data.append(value)

    full_data = data[3:]  # remove title of excel file
    return full_data


def process_lst(lst_in: list) -> dict:
    dict_output = {}
    for i in range(len(lst_in) // 3):
        dict_output[(lst_in[3 * i]).lower()] = [lst_in[3 * i + 1], int(lst_in[3 * i + 2])]

    return dict_output


def get_data_from_excel(excel_file_name: str) -> dict:
    return process_lst(read_data_from_excel(excel_file_name))


def write_data_to_excel(excel_file_name_out, data: dict):
    path_exel_file = path.join(getcwd(), "src", excel_file_name_out)
    sheet_name = "Sheet1"
    wb = openex(path_exel_file)
    sheet_obj = wb[sheet_name]
    # +2 because index from 0 and ignore title
    for index, key in enumerate(data):
        sheet_obj.cell(column=1, row=index + 2, value=key)
        sheet_obj.cell(column=2, row=index + 2, value=data.get(key)[0])
        sheet_obj.cell(column=3, row=index + 2, value=data.get(key)[1])

    wb.save(path_exel_file)
    wb.close()

def encrypt_data(key: str, data: dict):
    # encrypt data
    encrypt_obj = newCipher(key)
    output = {}
    for key in data:
        output[encrypt_obj.encrypt(key.encode()).hex()] = [encrypt_obj.encrypt(data.get(key)[0].encode()).hex(),
                                                           data.get(key)[1]]
    # output["Flag"] = ["Encrypt", h_obj.hash_str(key).hex()]
    return output

def decrypt_data(key: str, data: dict):
    if check_passwd(key):
        decrypt_obj = newCipher(key)
        output = {}
        for key in data:
            output[decrypt_obj.decrypt(bytes.fromhex(key)).decode()] = [
                decrypt_obj.decrypt(bytes.fromhex(data.get(key)[0])).decode(),
                data.get(key)[1]]
        return output
    else:
        return "ERROR"

def check_passwd(key: str):
    h_obj = newHash()
    with open(passwd_path, "rb") as f:
        data = f.read()
    hash_from_file = data.split(b'\n')[0]
    # status_encrypt = data.split(b'\n')[1]
    return hash_from_file == h_obj.hash_str(key)

def is_encrypted():
    with open(passwd_path, "rb") as f:
        data = f.read()
    status_encrypt = data.split(b'\n')[1]
    return status_encrypt == b'Encrypted'

def is_decrypted():
    with open(passwd_path, "rb") as f:
        data = f.read()
    status_encrypt = data.split(b'\n')[1]
    return status_encrypt == b'Decrypted'

def encrypt_excel(key: str, file_name: str):
    write_data_to_excel(file_name, encrypt_data(key, get_data_from_excel(file_name)))
    # write hash passwd to file
    h_obj = newHash()
    with open(passwd_path, "wb") as f:
        f.write(h_obj.hash_str(key))
        f.write(b'\nEncrypted')

def decrypt_excel(key: str, file_name: str):
    write_data_to_excel(file_name, decrypt_data(key, get_data_from_excel(file_name)))
    # write flag to file
    h_obj = newHash()
    with open(passwd_path, "wb") as f:
        f.write(h_obj.hash_str(key))
        f.write(b'\nDecrypted')

def read_and_decrypt_data(key: str, file_name: str):
    return decrypt_data(key, get_data_from_excel(file_name))

def encrypt_and_write(file_name: str, key: str, data: dict):
    write_data_to_excel(file_name, encrypt_data(key, data))
    h_obj = newHash()
    with open(passwd_path, "wb") as f:
        f.write(h_obj.hash_str(key))
        f.write(b'\nEncrypted')




if __name__ == '__main__':
    main()